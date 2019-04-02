import openpyxl
from io import BytesIO
from xlrd import open_workbook, xldate_as_tuple
from datetime import datetime


def read_by_openpyxl(filename=None, file_contents=None, title={}):
    '''
    title like this:
    {
        'GUID': 0, 
        '标题': 0, 
        'URL': 0, 
        '发布时间': 0, 
        '发布媒体': 0, 
        '风险程度': 0, 
        '地域': 0, 
        '类别': 0, 
    }
    '''
    # xlsx_book = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
    xlsx_book = openpyxl.load_workbook(filename=filename, read_only=True)
    sheet = xlsx_book.active
    rows = sheet.rows

    data = []
    for index, row in enumerate(rows):
        if index == 0:
            line = [cell.value for cell in row]
            for k in title.keys():
                title[k] = line.index(k)
        else:
            data2 = {}
            for k, v in title.items():
                data2[k] = sheet.cell(row=index, column=v).value
            data.append(data2)

    return data

def read_by_xlrd(filename=None, file_contents=None, title={}):
    '''
    title like this:
    {
        'GUID': 0, 
        '标题': 0, 
        'URL': 0, 
        '发布时间': 0, 
        '发布媒体': 0, 
        '风险程度': 0, 
        '地域': 0, 
        '类别': 0, 
    }
    '''
    wb = open_workbook(filename=filename, file_contents=file_contents)
    sheet = wb.sheet_by_index(0)
    rows = sheet.nrows
    cols = sheet.ncols

    rvalues = sheet.row_values(0)
    for k in title.keys():
        title[k] = rvalues.index(k)

    # DATA FORMAT
    def format(i, j):
        ctype = sheet.cell(i, j).ctype  # 表格的数据类型
        cell = sheet.cell_value(i, j)
        if ctype == 2 and cell % 1 == 0:  # 如果是整形
            cell = int(cell)
        elif ctype == 3: # 转成datetime对象
            cell = datetime(*xldate_as_tuple(cell, 0))
        elif ctype == 4:
            cell = True if cell == 1 else False

        return cell

    data = []
    for i in range(1, rows):
        data2 = {}
        for k, v in title.items():
            data2[k] = format(i, v)
        data.append(data2)

    return data

def write_by_openpyxl(filename, data, enterprises):
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet()

    # now we'll fill it with 100 rows x 200 columns
    for d in data:
        ws.append(d)

    # 导出企业：
    if enterprises:
        en = wb.create_sheet('企业')
        for e in enterprises:
            en.append(e)

    # save the file
    wb.save(filename)
