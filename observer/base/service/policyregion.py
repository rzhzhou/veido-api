import datetime
import openpyxl
from io import BytesIO

from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count, Q, F
from observer.apps.hqi.models import Policy, Area, PolicyArticle
from observer.base.service.abstract import Abstract
from observer.utils.excel import (read_by_openpyxl, write_by_openpyxl, )

# 区域
class PolicyAreaData(Abstract):

    def __init__(self, params={}):
        super(PolicyAreaData, self).__init__(params)

    def get_all(self):
        policydata = []
        cond = {
            'id': getattr(self, 'areas', None),
            'level': getattr(self, 'level', None),
        }
        args = dict([k, v] for k, v in cond.items() if v)
        areas = Area.objects.using('hqi').filter(policyarticle__policy__category='区域政策',**args).annotate(num_policies=Count('policyarticle'))
        for area in areas:
            if area.num_policies:
                level = Area.objects.using('hqi').filter(id=area.id).values_list('level',flat=True)[0]
                policies = Area.objects.using('hqi').filter(policyarticle__policy__category='区域政策',id=area.id).values_list('policyarticle__id',flat=True)
                fields = {'title','url'}
                articles = PolicyArticle.objects.using('hqi').filter(id__in=policies).values(*fields)
                for x in map(lambda r: {'url': r['url'],'title': r['title'],},articles):
                    articletitle = x['title']
                    articleurl = x['url']
                    nums = PolicyArticle.objects.using('hqi').filter(url=articleurl).count()
                    a = {'area':area.name,
                        'area__id':area.id,
                        'total':nums,
                        'level':level,
                        'articletitle':articletitle,
                        'articleurl':articleurl,
                        }
                    if a not in policydata:
                        policydata.append(a)


        return policydata


class PolicyAreaTotalData(Abstract):

    def __init__(self, params={}):
        super(PolicyAreaTotalData, self).__init__(params)

    def get_all(self):
        fields = ('policy__id', 'policy__name','policy__detail','pubtime')
        url = getattr(self, 'url', None)
        queryset = PolicyArticle.objects.using('hqi').filter(url=url)

        return queryset.values(*fields)


class PolicyAreaAdd(Abstract):

    def __init__(self, user, params={}):
        super(PolicyAreaAdd, self).__init__(params)
        self.user = user

    def add(self):
        title = getattr(self,'title', '')
        url = getattr(self, 'url', '')
        pubtime = getattr(self, 'pubtime', '')
        areas = getattr(self, 'areas', '')
        name = getattr(self, 'name', '')
        detail = getattr(self, 'detail', '')

        if not PolicyArticle.objects.using('hqi').filter(url=url, title=title).exists():
            policy = Policy(
                category = '区域政策',
                name = name,
                detail = detail,
            )
            policy.save(using='hqi')
            policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
            policyarticle =PolicyArticle(
                title = title,
                url = url,
                pubtime = pubtime,
                policy_id= policyid,
            )
            policyarticle.save(using='hqi')
            area_id = Area.objects.using('hqi').filter(name=areas).values_list('id', flat=True)
            areas = Area.objects.using('hqi').filter(id__in=area_id)
            policyarticle.areas.add(*areas)
            policyarticle.save(using='hqi')

        else:
            if not Policy.objects.using('hqi').filter(category='区域政策',name=name,detail=detail).exists():
                policy = Policy(
                    category = '区域政策',
                    name = name,
                    detail = detail,
                )
                policy.save(using='hqi')
                policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
                policyarticle =PolicyArticle(
                    title = title,
                    url = url,
                    pubtime = pubtime,
                    policy_id= policyid,
                )
                policyarticle.save(using='hqi')
                area_id = Area.objects.using('hqi').filter(name=areas).values_list('id', flat=True)
                areas = Area.objects.using('hqi').filter(id__in=area_id)
                policyarticle.areas.add(*areas)
                policyarticle.save(using='hqi')


# 民营
class PolicyPrivateData(Abstract):

    def __init__(self, params={}):
        super(PolicyPrivateData, self).__init__(params)

    def get_all(self):
        policydata = []
        cond = {
            'id': getattr(self, 'areas', None),
            'level': getattr(self, 'level', None),
        }
        args = dict([k, v] for k, v in cond.items() if v)
        areas = Area.objects.using('hqi').filter(policyarticle__policy__category='民营政策',**args).annotate(num_policies=Count('policyarticle'))
        for area in areas:
            if area.num_policies:
                level = Area.objects.using('hqi').filter(id=area.id).values_list('level',flat=True)[0]
                policies = Area.objects.using('hqi').filter(policyarticle__policy__category='民营政策',id=area.id).values_list('policyarticle__id',flat=True)
                fields = {'title','url'}
                articles = PolicyArticle.objects.using('hqi').filter(id__in=policies).values(*fields)
                for x in map(lambda r: {'url': r['url'],'title': r['title'],},articles):
                    articletitle = x['title']
                    articleurl = x['url']
                    nums = PolicyArticle.objects.using('hqi').filter(url=articleurl).count()
                    a = {'area':area.name,
                        'area__id':area.id,
                        'total':nums,
                        'level':level,
                        'articletitle':articletitle,
                        'articleurl':articleurl,
                        }
                    if a not in policydata:
                        policydata.append(a)


        return policydata


class PolicPrivatelTotalData(Abstract):

    def __init__(self, params={}):
        super(PolicPrivatelTotalData, self).__init__(params)

    def get_all(self):
        fields = ('policy__id', 'policy__name','policy__detail','pubtime')
        url = getattr(self, 'url', None)
        queryset = PolicyArticle.objects.using('hqi').filter(url=url)

        return queryset.values(*fields)


class PolicPrivatelAdd(Abstract):

    def __init__(self, user, params={}):
        super(PolicPrivatelAdd, self).__init__(params)
        self.user = user

    def add(self):
        title = getattr(self,'title', '')
        url = getattr(self, 'url', '')
        pubtime = getattr(self, 'pubtime', '')
        areas = getattr(self, 'areas', '')
        name = getattr(self, 'name', '')
        detail = getattr(self, 'detail', '')

        if not PolicyArticle.objects.using('hqi').filter(url=url, title=title).exists():
            policy = Policy(
                category = '民营政策',
                name = name,
                detail = detail,
            )
            policy.save(using='hqi')
            policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
            policyarticle =PolicyArticle(
                title = title,
                url = url,
                pubtime = pubtime,
                policy_id= policyid,
            )
            policyarticle.save(using='hqi')
            area_id = Area.objects.using('hqi').filter(name=areas).values_list('id', flat=True)
            areas = Area.objects.using('hqi').filter(id__in=area_id)
            policyarticle.areas.add(*areas)
            policyarticle.save(using='hqi')

        else:
            if not Policy.objects.using('hqi').filter(category='民营政策',name=name,detail=detail).exists():
                policy = Policy(
                    category = '民营政策',
                    name = name,
                    detail = detail,
                )
                policy.save(using='hqi')
                policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
                policyarticle =PolicyArticle(
                    title = title,
                    url = url,
                    pubtime = pubtime,
                    policy_id= policyid,
                )
                policyarticle.save(using='hqi')
                area_id = Area.objects.using('hqi').filter(name=areas).values_list('id', flat=True)
                areas = Area.objects.using('hqi').filter(id__in=area_id)
                policyarticle.areas.add(*areas)
                policyarticle.save(using='hqi')


# 产业
class PolicyIndustryData(Abstract):

    def __init__(self, params={}):
        super(PolicyIndustryData, self).__init__(params)

    def get_all(self):
        policydata = []
        policies = []
        policyarticles = []
        cond = {
            'areas__name': getattr(self, 'name', None),
            'policy__industry': getattr(self, 'industry', None),
            'areas__level': getattr(self, 'level', None),
        }
        args = dict([k, v] for k, v in cond.items() if v)
        areas = Area.objects.using('hqi').filter(policyarticle__policy__category='产业政策').annotate(num_policyarticles=Count('policyarticle'))
        for area in areas:
            if area.num_policyarticles:
                fields =('url','title','policy__industry','areas__name','areas__id','areas__level')
                policyarticle = PolicyArticle.objects.using('hqi').filter(**args,areas__id=area.id,policy__category='产业政策').values(*fields)
                for x in range(len(policyarticle)):
                    policyarticle_url = policyarticle[x]['url']
                    policyarticle_title = policyarticle[x]['title']
                    policy__industry = policyarticle[x]['policy__industry']
                    areas__name = policyarticle[x]['areas__name']
                    areas__id = policyarticle[x]['areas__id']
                    areas__level = policyarticle[x]['areas__level']
                    totals = PolicyArticle.objects.using('hqi').filter(url=policyarticle_url).count()
                    a = {'area':areas__name,
                        'area__id':areas__id,
                        'total':totals,
                        'level':areas__level,
                        'articletitle':policyarticle_title,
                        'articleurl':policyarticle_url,
                        'industry': policy__industry,
                        }
                    if a not in policydata:
                        policydata.append(a)
        return policydata


class PolicyIndustryTotalData(Abstract):

    def __init__(self, params={}):
        super(PolicyIndustryTotalData, self).__init__(params)

    def get_all(self):
        fields = ('policy__id', 'policy__name','policy__detail','pubtime')
        url = getattr(self, 'url', None)
        queryset = PolicyArticle.objects.using('hqi').filter(url=url,policy__category='产业政策')

        return queryset.values(*fields)


class PolicyIndustryAdd(Abstract):

    def __init__(self, user, params={}):
        super(PolicyIndustryAdd, self).__init__(params)
        self.user = user

    def add(self):
        title = getattr(self,'title', '')
        url = getattr(self, 'url', '')
        pubtime = getattr(self, 'pubtime', '')
        areas = getattr(self, 'areas', '')
        name = getattr(self, 'name', '')
        detail = getattr(self, 'detail', '')
        industrys = getattr(self,'industrys','')

        if not PolicyArticle.objects.using('hqi').filter(url=url, title=title).exists():
            policy = Policy(
                category = '产业政策',
                industry = industrys,
                name = name,
                detail = detail,
            )
            policy.save(using='hqi')
            policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
            policyarticle =PolicyArticle(
                title = title,
                url = url,
                pubtime = pubtime,
                policy_id= policyid,
            )
            policyarticle.save(using='hqi')
            policyarticle.areas.add(areas)
            policyarticle.save(using='hqi')

        else:
            if not Policy.objects.using('hqi').filter(category='产业政策',name=name,detail=detail).exists():
                policy = Policy(
                    category = '产业政策',
                    industry = industrys,
                    name = name,
                    detail = detail,
                )
                policy.save(using='hqi')
                policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
                policyarticle =PolicyArticle(
                    title = title,
                    url = url,
                    pubtime = pubtime,
                    policy_id= policyid,
                )
                policyarticle.save(using='hqi')
                policyarticle.areas.add(areas)
                policyarticle.save(using='hqi')


#导入
class PolicyDataUpload(Abstract):

    def __init__(self, user):
        self.user = user

    def upload(self, filename, file_obj):
        #Model weight
        model = {'地域': 0,'政策类别': 0,  'URL':0 , '标题':0 ,'政策': 0, '详细':0 , '时间':0}
        #sheet value
        sv = lambda x, y, z : z.cell(row=x, column=y).value
        #date format
        def date_format(df):
            try:
                return openpyxl.utils.datetime.from_excel(df)
            except Exception:
                try:
                    return str_to_date(df)
                except Exception:
                    return df

        try:
            xlsx_book = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
            sheet = xlsx_book.active
            rows = sheet.rows
        except Exception as e:
            return {
                    'status': 0,
                    'message': '操作失败！请检查文件是否有误。详细错误信息：%s！' % e
                }

        total = 0
        dupli = 0

        for i, row in enumerate(rows):
            i += 1
            if i == 1:
                line = [cell.value for cell in row]
                for k in model.keys():
                    model[k] = line.index(k) + 1
            else:
                try:
                    # 政策类别
                    category = sv(i, model['政策类别'], sheet)
                    if category == 'None':
                        continue

                    # 政策
                    name = sv(i, model['政策'], sheet)
                    if name == 'None':
                        continue

                    # URL
                    url = sv(i, model['URL'], sheet)

                    # 时间
                    pubtime = sv(i, model['时间'], sheet)

                    # 标题
                    title = sv(i, model['标题'], sheet)

                    # 详细
                    detail = sv(i, model['详细'], sheet)

                    try:
                        # 地域
                        area = sv(i, model['地域'], sheet)
                        if area == 'None':
                            continue
                        area_id = Area.objects.using('hqi').get(name = area).id


                    except Exception as e:
                        return {
                            'status': 0,
                            'message': '地域不存在！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
                        }


                    total += 1

                    old_policyarticle = PolicyArticle.objects.using('hqi').filter(url=url,policy__name=name)
                    if not old_policyarticle.exists():
                        policy = Policy(
                            category = category,
                            name = name,
                            detail = detail,
                        )
                        policy.save(using='hqi')
                        policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
                        policyarticle =PolicyArticle(
                            title = title,
                            url = url,
                            pubtime = pubtime,
                            policy_id= policyid,
                        )
                        policyarticle.save(using='hqi')
                        policyarticle.areas.add(area_id)
                        policyarticle.save(using='hqi')
                        dupli += 1


                except Exception as e:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
                    }
        return {
                'status': 1,
                'message': '操作成功！共处理%s条数据，新增数据%s条，更新数据%s条！' % (total, total - dupli, dupli, )
                }


class PolicyIndustryDataUpload(Abstract):

    def __init__(self, user):
        self.user = user

    def upload(self, filename, file_obj):
        #Model weight
        model = {'地域': 0,'政策类别': 0,'产业类别': 0,  'URL':0 , '标题':0 ,'政策': 0, '详细':0 , '时间':0}

        #sheet value
        sv = lambda x, y, z : z.cell(row=x, column=y).value
        #date format
        def date_format(df):
            try:
                return openpyxl.utils.datetime.from_excel(df)
            except Exception:
                try:
                    return str_to_date(df)
                except Exception:
                    return df

        try:
            xlsx_book = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
            sheet = xlsx_book.active
            rows = sheet.rows
        except Exception as e:
            return {
                    'status': 0,
                    'message': '操作失败！请检查文件是否有误。详细错误信息：%s！' % e
                }

        total = 0
        dupli = 0

        for i, row in enumerate(rows):
            i += 1
            if i == 1:
                line = [cell.value for cell in row]
                for k in model.keys():
                    model[k] = line.index(k) + 1
            else:
                try:
                    # 政策类别
                    category = sv(i, model['政策类别'], sheet)
                    if category == 'None':
                        continue

                    # 产业类别
                    industry = sv(i, model['产业类别'], sheet)
                    if industry == 'None':
                        continue

                    # 政策
                    name = sv(i, model['政策'], sheet)
                    if name == 'None':
                        continue

                    # URL
                    url = sv(i, model['URL'], sheet)

                    # 时间
                    pubtime = sv(i, model['时间'], sheet)

                    # 标题
                    title = sv(i, model['标题'], sheet)

                    # 详细
                    detail = sv(i, model['详细'], sheet)

                    try:
                        # 地域
                        area = sv(i, model['地域'], sheet)
                        if area == 'None':
                            continue
                        area_id = Area.objects.using('hqi').get(name = area).id

                    except Exception as e:
                        return {
                            'status': 0,
                            'message': '地域不存在！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
                        }


                    total += 1

                    old_policyarticle = PolicyArticle.objects.using('hqi').filter(url=url,policy__name=name)
                    if not old_policyarticle.exists():
                        policy = Policy(
                            category = category,
                            industry = industry,
                            name = name,
                            detail = detail,
                        )
                        policy.save(using='hqi')
                        policyid = Policy.objects.using('hqi').filter(name=name,detail=detail).values_list('id',flat=True)[0]
                        policyarticle =PolicyArticle(
                            title = title,
                            url = url,
                            pubtime = pubtime,
                            policy_id= policyid,
                        )
                        policyarticle.save(using='hqi')
                        policyarticle.areas.add(area_id)
                        policyarticle.save(using='hqi')
                        dupli += 1


                except Exception as e:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
                    }
        return {
                    'status': 1,
                    'message': '操作成功！共处理%s条数据，新增数据%s条，更新数据%s条！' % (total, total - dupli, dupli, )
                }
