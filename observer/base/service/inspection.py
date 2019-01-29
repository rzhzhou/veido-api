from io import BytesIO

import threading

import jieba
import jieba.posseg as pseg

import openpyxl
from django.core.exceptions import ObjectDoesNotExist
from django.contrib.auth.models import Group, User
from django.db.models import Q

from observer.base.models import (MajorIndustry, AliasIndustry, UserArea, Area, Enterprise, Industry,
                                  Inspection, IndustryProducts)
from observer.base.service.abstract import Abstract
from observer.base.service.base import (alias_industry, get_major_industry, area,
                                        enterprise_area_name, enterprise_name,
                                        enterprise_unitem, industry_number,
                                        qualitied, involve_local)
from observer.utils.date_format import date_format, get_months, str_to_date
from observer.utils.excel import read_by_openpyxl, write_by_openpyxl
from observer.utils.str_format import str_to_md5str
from observer.utils.crawler.enterprise_crawler import crawler


class InspectionData(Abstract):

    def __init__(self, params):
        super(InspectionData, self).__init__(params)

    def get_all(self):

        fields = ('id', 'industry__name', 'area__name', 'title', 'url', 'pubtime', 'source',
                  'qualitied', 'unqualitied_patch', 'qualitied_patch',
                  'inspect_patch', 'category', 'level', 'industry',
                  'area', 'product_name', 'status', 'origin_product')

        cond = {
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
            'qualitied__gte': getattr(self, 'qualitied_gte', None),
            'qualitied__lt': getattr(self, 'qualitied_lt', None),
            'category__contains': getattr(self, 'category', None),
            'source__contains': getattr(self, 'source', None),
            'level': getattr(self, 'level', None),
            'industry': getattr(self, 'industry', None),
            'area': getattr(self, 'area', None),
            'product_name__icontains': getattr(self, 'product_name', None),
            'status__in': getattr(self, 'status').split(',') if getattr(self, 'status', None) else None
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Inspection.objects.filter(**args).values(*fields).order_by('-pubtime')

        return queryset


class EnterpriseDataEdit(Abstract):

    def __init__(self, params={}):
        super(EnterpriseDataEdit, self).__init__(params)

    def edit(self, cid):
        edit_id = cid
        name = getattr(self, 'name', '')
        unitem = getattr(self, 'unitem', '')
        area_id = getattr(self, 'area_id', '')

        if not name or not unitem or not area_id:
            return 400

        enterprise = Enterprise.objects.get(id=edit_id)
        enterprise.name = name
        enterprise.unitem = unitem
        enterprise.area_id = area_id
        enterprise.save()

        return 200


class EnterpriseDataDelete(Abstract):

    def __init__(self, user):
        self.user = user

    def delete(self, cid):
        del_ids = cid.split(",")

        Enterprise.objects.filter(id__in=del_ids).delete()

        return 200


class EnterpriseDataAudit(Abstract):
    def __init__(self, params={}):
        super(EnterpriseDataAudit, self).__init__(params)

    def edit(self, cid):
        audit_ids = cid.split(",")

        Enterprise.objects.filter(id__in=audit_ids).update(status=1)

        return 200


class EnterpriseDataUnqualified(Abstract):
    def __init__(self, params):
        super(EnterpriseDataUnqualified, self).__init__(params)

    def get_all(self):

        fields = ('inspection__industry', 'inspection__industry__name',
                  'inspection__product_name', 'inspection__source', 'name',
                  'area', 'area__name', 'unitem', 'inspection__pubtime')

        cond = {
            'inspection__product_name__contains': getattr(self, 'productName', None),
            'name__contains': getattr(self, 'enterpriseName', None),
            'area': getattr(self, 'Area', None),
            'inspection__industry': getattr(self, 'industry', None),
            'inspection__pubtime__gte': getattr(self, 'starttime', None),
            'inspection__pubtime__lte': getattr(self, 'endtime', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Enterprise.objects.filter(**args, status=1).values(*fields).order_by('-inspection__pubtime')

        return queryset


class EnterpriseData(Abstract):

    def __init__(self, params):
        super(EnterpriseData, self).__init__(params)

    def get_all(self):
        fields = ('id', 'name', 'unitem', 'area_id', 'status' )

        cond = {
            'name__contains': getattr(self, 'name', None),
            'unitem__contains': getattr(self, 'unitem', None),
            'area_id': getattr(self, 'area', None),
            'status': getattr(self, 'status', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Enterprise.objects.filter(**args).values(*fields).order_by('-id')

        return queryset

    def get_by_id(self, eid):

        queryset = Enterprise.objects.filter(inspection=eid)

        return queryset


class InspectionDataAdd(Abstract):

    def __init__(self, user, params={}):
        super(InspectionDataAdd, self).__init__(params)
        self.user = user

    def add(self):
        # qualification rate
        def qr(x, y): return float(0) if not x else float(x) / float(y)

        title = getattr(self, 'title', '')
        url = getattr(self, 'url', '')
        pubtime = getattr(self, 'pubtime', '')
        source = getattr(self, 'source', '')

        inspect_patch = getattr(self, 'inspect_patch', 0)
        qualitied_patch = getattr(self, 'qualitied_patch', 0)
        unqualitied_patch = getattr(self, 'unqualitied_patch', 0)

        category = getattr(self, 'category', '')
        level = getattr(self, 'level', '')
        if level == '市':
            new_level = 0
        elif level == '省':
            new_level = 1
        elif level == '国':
            new_level = 2

        area_id = getattr(self, 'area_id', '')
        product_id = getattr(self, 'product', '')

        product_name = IndustryProducts.objects.filter(id=product_id)[0].name
        industry_id = IndustryProducts.objects.filter(id=product_id)[0].industry_id

        if not url or not pubtime or not source or not inspect_patch or not qualitied_patch or not unqualitied_patch or not level or not product_name or not area_id:
            return 400

        Inspection(
            title=title,
            url=url,
            pubtime=pubtime,
            source=source,
            qualitied=qr(qualitied_patch, inspect_patch),
            inspect_patch=inspect_patch,
            qualitied_patch=qualitied_patch,
            unqualitied_patch=unqualitied_patch,
            category=category,
            level=new_level,
            product_name=product_name,
            industry_id=industry_id,
            area_id=area_id,
            status=0,
        ).save()

        return 200


class InspectionDataEdit(Abstract):

    def __init__(self, user, params={}):
        super(InspectionDataEdit, self).__init__(params)
        self.user = user

    def edit(self, cid):
        # qualification rate
        def qr(x, y): return float(0) if not x else float(x) / float(y)

        edit_id = cid
        url = getattr(self, 'url', '')
        pubtime = getattr(self, 'pubtime', '')
        source = getattr(self, 'source', '')

        inspect_patch = getattr(self, 'inspect_patch', 0)
        qualitied_patch = getattr(self, 'qualitied_patch', 0)
        unqualitied_patch = getattr(self, 'unqualitied_patch', 0)

        category = getattr(self, 'category', '')
        level = getattr(self, 'level', '')
        area_id = getattr(self, 'area_id', '')
        product_id = getattr(self, 'product', '')

        product_name = IndustryProducts.objects.filter(id=product_id)[0].name
        industry_id = IndustryProducts.objects.filter(id=product_id)[0].industry_id

        if not url or not pubtime or not source or not inspect_patch or not qualitied_patch or not unqualitied_patch or not level or not industry_id or not area_id:
            return 400

        inspection = Inspection.objects.get(id=edit_id)
        inspection.url = url
        inspection.pubtime = pubtime
        inspection.source = source
        inspection.qualitied = qr(qualitied_patch, inspect_patch)
        inspection.inspect_patch = inspect_patch
        inspection.qualitied_patch = qualitied_patch
        inspection.unqualitied_patch = unqualitied_patch
        inspection.category = category
        inspection.level = level
        inspection.product_name = product_name
        inspection.industry_id = industry_id
        inspection.area_id = area_id
        inspection.save()

        return 200


class InspectionDataDelete(Abstract):

    def __init__(self, user):
        self.user = user

    def delete(self, cid):
        del_ids = cid.split(",")

        Inspection.objects.filter(id__in=del_ids).delete()

        return 200


class InspectionDataUpload(Abstract):

    def __init__(self, user):
        self.user = user

    def upload(self, filename, file_obj):
        jieba.load_userdict('observer/utils/dictionary.txt')

        # ModelWeight
        model = {'标题': 0, '链接': 0, '发布日期': 0, '抽查类别': 0, '抽查等级': 0, '抽检单位': 0,
                 '地域': 0, '产品名称': 0, '抽查批次': 0, '合格批次': 0, '不合格批次': 0}
        # qualification rate

        def qr(x, y): return float(0) if not x else float(x) / float(y)
        # sheet values

        def sv(x, y, z): return z.cell(row=x, column=y).value
        # date format

        def date_format(df):
            try:
                return openpyxl.utils.datetime.from_excel(df)
            except Exception:
                try:
                    return str_to_date(df)
                except Exception:
                    return df

        try:
            xlsx_book = openpyxl.load_workbook(
                BytesIO(file_obj.read()), read_only=True)
            sheet = xlsx_book.active
            rows = sheet.rows
        except Exception as e:
            return {
                'status': 0,
                'message': '操作失败！请检查文件是否有误。详细错误信息：%s！' % e
            }

        total = 0
        data_list = []

        for i, row in enumerate(rows):
            i += 1
            if i == 1:
                line = [cell.value for cell in row]
                for k in model.keys():
                    model[k] = line.index(k) + 1
            else:
                title = sv(i, model['标题'], sheet)
                url = sv(i, model['链接'], sheet)

                if not url:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行"链接"有误！' % (i + 1, )
                    }

                pubtime = date_format(sv(i, model['发布日期'], sheet))
                if not pubtime:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行"时间格式"有误！' % (i + 1, )
                    }

                category = sv(i, model['抽查类别'], sheet)

                level = sv(i, model['抽查等级'], sheet)
                if level == '市':
                    new_level = 0
                elif level == '省':
                    new_level = 1
                elif level == '国':
                    new_level = 2
                else:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行"抽查等级"有误！' % (i + 1, )
                    }

                source = sv(i, model['抽检单位'], sheet)
                area_name = sv(i, model['地域'], sheet)
                product_undisposed = sv(i, model['产品名称'], sheet)
                origin_product = sv(i, model['产品名称'], sheet) # 导入产品名称
                inspect_patch = sv(i, model['抽查批次'], sheet)
                qualitied_patch = sv(i, model['合格批次'], sheet)
                unqualitied_patch = sv(i, model['不合格批次'], sheet)

                total += 1

                # 处理产品名称
                speech = pseg.cut(product_undisposed)
                j = 0
                k = 0
                product_name = []
                for word, flag in speech: # speech: 电动车充电器 和 电瓶, 网售 玩具, 质量 维权 商品
                    if flag == 'findproduct':
                        j += 1
                        if j >= 2:
                            continue
                        product_name.append(word)

                if j == 0:
                    word = 'None'
                    k += 1
                    if k >= 2:
                        continue
                    product_name.append(word)

                j = 0
                k = 0

                # 处理抽检地域
                area = Area.objects.filter(name=area_name)
                if not area.exists():
                    return {
                        'status': 0,
                        'message': '操作失败！Excel第%s行，地域（%s）不存在！' % (i, area_name, )
                    }

                industry_id = IndustryProducts.objects.filter(name=product_name[0])[0].industry_id

                bulk_inspection = Inspection(
                    title=title,
                    url=url,
                    pubtime=pubtime,
                    source=source,
                    qualitied=qr(qualitied_patch, inspect_patch),
                    unqualitied_patch=unqualitied_patch,
                    qualitied_patch=qualitied_patch,
                    inspect_patch=inspect_patch,
                    category='' if not category else category,
                    level=new_level,
                    industry_id=industry_id,
                    product_name = product_name[0],
                    origin_product=origin_product,
                    area_id=area[0].id,
                    status=0,
                )
                data_list.append(bulk_inspection)

        Inspection.objects.bulk_create(data_list)

        return {
            'status': 1,
            'message': '操作成功！共处理%s条数据！' % total
        }


class InspectionDataUnEnterpriseUpload(Abstract):

    def __init__(self, user):
        self.user = user

#     def upload(self, filename, file_obj):
#         # ModelWeight
#         model = {'链接': 0, '产品名称': 0,
#                  '不合格企业': 0, '不合格企业地域': 0, '不合格项': 0}
#         # sheet values

#         def sv(x, y, z): return z.cell(row=x, column=y).value

#         try:
#             xlsx_book = openpyxl.load_workbook(
#                 BytesIO(file_obj.read()), read_only=True)
#             sheet = xlsx_book.active
#             rows = sheet.rows
#         except Exception as e:
#             return {
#                 'status': 0,
#                 'message': '操作失败！请检查文件是否有误。详细错误信息：%s！' % e
#             }

#         total = 0
#         dupli = 0

#         for i, row in enumerate(rows):
#             i += 1
#             if i == 1:
#                 line = [cell.value for cell in row]
#                 for k in model.keys():
#                     model[k] = line.index(k) + 1
#             else:
#                 try:
#                     url = sv(i, model['链接'], sheet)

#                     if not url:
#                         continue

#                     product_name = sv(i, model['产品名称'], sheet)
#                     unenterprise = sv(i, model['不合格企业'], sheet)
#                     unenterprise_area = sv(i, model['不合格企业地域'], sheet)
#                     unitem = sv(i, model['不合格项'], sheet)

#                     total += 1

#                     # 处理抽检地域
#                     area = Area.objects.filter(name=unenterprise_area)
#                     if not area.exists():
#                         return {
#                             'status': 0,
#                             'message': '操作失败！Excel第%s行，地域（%s）不存在！' % (i, unenterprise_area, )
#                         }

#                     area_id = area[0].id

#                     guid = str_to_md5str('{0}{1}'.format(url, product_name))

#                     # 处理不合格企业信息
#                     enterprise = Enterprise.objects.filter(
#                         name=unenterprise, area_id=area_id)
#                     if not enterprise.exists():
#                         Enterprise(
#                             name=unenterprise,
#                             unitem=unitem,
#                             area_id=area_id,
#                         ).save()

#                     enterprise_id = Enterprise.objects.filter(
#                         name=unenterprise, area_id=area_id)[0].id
#                     inspection_enterprise = InspectionEnterprise.objects.filter(
#                         inspection_id=guid, enterprise_id=enterprise_id)
#                     if not inspection_enterprise.exists():
#                         InspectionEnterprise(
#                             inspection_id=guid,
#                             enterprise_id=enterprise_id,
#                         ).save()
#                     else:
#                         dupli += 1

#                 except Exception as e:
#                     return {
#                         'status': 0,
#                         'message': '操作失败！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
#                     }

#         return {
#             'status': 1,
#             'message': '操作成功！共处理%s条数据，成功导入%s条数据，重复数据%s条！' % (total, total - dupli, dupli, )
#         }


class InspectionDataExport(Abstract):

    def __init__(self, user):
        self.user = user

    def export(self):
        filename = "inspections.xlsx"

        # process data
        data = [
            ['标题', '链接', '发布日期', '抽查类别', '抽查等级', '抽检单位', '地域',
                '行业编号', '产品名称', '不合格企业', '不合格企业地域', '不合格项', '合格率', ],
        ]
        months = get_months()[-1::][0]
        start = months[0].strftime('%Y-%m-%d')
        end = months[1].strftime('%Y-%m-%d')

        queryset = Inspection.objects.filter(pubtime__gte=start, pubtime__lte=end).values(
            'id', 'title', 'url', 'pubtime', 'category', 'level', 'source', 'area__name', 'industry_id', 'industry__name', 'qualitied', 'enterprises__name',
            'enterprises__area__name', 'enterprises__unitem')

        for q in queryset:
            data.append([
                q['title'],
                q['url'],
                date_format(q['pubtime'], '%Y-%m-%d'),
                q['category'],
                q['level'],
                q['source'],
                q['area__name'],
                q['industry_id'],
                q['industry__name'],
                q['enterprises__name'],
                q['enterprises__area__name'],
                q['enterprises__unitem'],
                qualitied(q['qualitied']),
            ])

        # write file
        write_by_openpyxl(filename, data)

        return open(filename, 'rb')


class InspectionDataCrawler(Abstract):

    def __init__(self, user, params={}):
        super(InspectionDataCrawler, self).__init__(params)
        self.user = user

    def edit(self, cid):
        edit_ids = cid
        urls = getattr(self, 'url', '')
        product_names = getattr(self, 'product', '')

        t = threading.Thread(target=crawler, args=(edit_ids, urls, product_names))
        t.start()

        return 200


class InspectionDataAudit(Abstract):

    def __init__(self, user, params={}):
        super(InspectionDataAudit, self).__init__(params)
        self.user = user

    def edit(self, cid):
        edit_ids = cid
        status = getattr(self, 'status', '')

        for ids in edit_ids.split(","):

            inspection = Inspection.objects.get(id=ids)
            inspection.status = status
            inspection.save()

        return 200


class InspectionDataSuzhou(Abstract):

    def __init__(self, params={}):
        super(InspectionDataSuzhou, self).__init__(params)

    def get_inspection_list(self, search_value):
        fields = ('id', 'industry_id', 'title', 'pubtime',
                  'url', 'level', 'area_id', 'source', 'qualitied',
                  'unqualitied_patch', 'qualitied_patch', 'inspect_patch',
                  'category', 'product_name', 'status' )

        args = {}
        if not search_value:
            queryset = Inspection.objects.filter(**args).values(*fields)
        else:
            queryset = Inspection.objects.filter(Q(source__contains=search_value) | Q(product_name__contains=search_value)).values(*fields)

        return queryset


class InspectionDataNation(Abstract):
    def __init__(self, params):
        super(InspectionDataNation, self).__init__(params)

    def get_all(self):
        fields = ('inspection__industry__name', 'unitem', 'area__name', 'inspection__pubtime')

        cond = {
            'inspection__level' : getattr(self, '', 2),
            'inspection__pubtime__year': getattr(self,'year',None),
            'inspection__pubtime__month': getattr(self,'month',None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Enterprise.objects.filter(**args).values(*fields)

        return queryset


class InspectionDataProAndCity(Abstract):
    def __init__(self, params):
        super(InspectionDataProAndCity, self).__init__(params)

    def get_all(self):
        fields = ('industry__name', 'qualitied', 'source', 'pubtime')

        cond = {
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Inspection.objects.filter(**args).values(*fields).order_by('qualitied')

        return queryset


class InspectionDataLocal(Abstract):

    def __init__(self, user, params):
        super(InspectionDataLocal, self).__init__(params)
        self.user = user

    def get_all(self):
        area_name = User.objects.filter(username=self.user).values('userarea__area__name')

        fields = ('inspection__industry__name', 'name', 'unitem','inspection__source','inspection__pubtime')

        cond = {
            'area__name': getattr(self, '', area_name[0]['userarea__area__name']),
            'inspection__pubtime__gte': getattr(self, 'starttime', None),
            'inspection__pubtime__lte': getattr(self, 'endtime', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Enterprise.objects.filter(**args).values(*fields)

        return queryset


class InspectionDataLocalExport(Abstract):

    def __init__(self, user, params):
        super(InspectionDataLocalExport, self).__init__(params)
        self.user = user

    def export(self):
        filename = "o.xlsx"

        # process data
        data = [
            ['序号', '抽查产品', '企业名称', '不合格项目', '抽检单位'],
        ]
  
        area_name = User.objects.filter(username=self.user).values('userarea__area__name')

        fields = ('inspection__industry__name', 'name', 'unitem', 'inspection__source', 'inspection__pubtime')

        cond = {
            'area__name': getattr(self, '', area_name[0]['userarea__area__name']),
            'inspection__pubtime__gte': getattr(self, 'starttime', None),
            'inspection__pubtime__lte': getattr(self, 'endtime', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Enterprise.objects.filter(**args).values(*fields)
        
        print(queryset)

        i = 1

        for q in queryset:
            data.append([
                i,
                q['inspection__industry__name'],
                q['name'],
                q['unitem'],
                q['inspection__source'],
            ])
            i = i+1
           

        # write file
        write_by_openpyxl(filename, data)

        return open(filename, 'rb')


class InspectionDataProAndCityExport(Abstract):

    def __init__(self, user, params):
        super(InspectionDataProAndCityExport, self).__init__(params)
        self.user = user

    def export(self):
        filename = "o.xlsx"

        # process data
        data = [
            ['序号','产品种类','合格率','抽检单位','发布日期'],
        ]

        fields = ('industry__name', 'qualitied', 'source', 'pubtime')

        cond = {
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Inspection.objects.filter(**args).values(*fields).order_by('qualitied')
  
        i = 1

        for q in queryset:
            data.append([
                i,
                q['industry__name'],
                str(round(float(q['qualitied'])*100,2))+'%',
                q['source'],
                q['pubtime'],
            ])
            i = i+1
           

        # write file
        write_by_openpyxl(filename, data)

        return open(filename, 'rb')


class InspectionDataNationExport(Abstract):

    def __init__(self, user, params):
        super(InspectionDataNationExport, self).__init__(params)
        self.user = user

    def export(self):
        filename = "o.xlsx"

        # process data
        data = [
            ['序号', '不合格产品种类','不合格项目', '是否涉及本地企业'],
        ]

        areaname = User.objects.filter(username=self.user).values('userarea__area__name')

        fields = ('inspection__industry__name', 'unitem', 'area__name', 'inspection__pubtime')

        cond = {
            'inspection__level':getattr(self, '', 2),
            'inspection__pubtime__year': getattr(self,'year',None),
            'inspection__pubtime__month': getattr(self,'month',None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Enterprise.objects.filter(**args).values(*fields).distinct()

        i = 1

        for q in queryset:
            data.append([
                i,
                q['inspection__industry__name'],
                q['unitem'],
                involve_local(areaname[0]['userarea__area__name'], q['area__name']),
            ])
            i = i+1
           

        # write file
        write_by_openpyxl(filename, data)

        return open(filename, 'rb')
