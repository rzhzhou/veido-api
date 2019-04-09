import openpyxl
from io import BytesIO
import threading

import datetime, copy
from django.core.exceptions import ObjectDoesNotExist
from django.db.models import Count, Q, F, Max, Min

from observer.base.models import(Area, Article, Category,
                                CorpusCategories, HarmIndicator, Harm,
                                HarmPeople, Events, EventsKeyword,
                                EventsMedia)
from django.contrib.auth.models import Group, User
from observer.base.service.abstract import Abstract
from observer.base.service.base import (areas, categories, )
from observer.utils.date_format import (date_format, str_to_date, get_months)
from observer.utils.str_format import str_to_md5str
from observer.utils.excel import (read_by_openpyxl, write_by_openpyxl, )
from observer.utils.crawler.news_crawler import newsCrawler
from observer.utils.crawler.baiDuAIP import BaiDuAI


class ArticleData(Abstract):

    def __init__(self, params, category):
        self.category = category
        super(ArticleData, self).__init__(params)

    def get_all(self):
        fields = ('id', 'url', 'title', 'source', 'pubtime', 'score', 'harm__id')

        cond = {
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
            'title__contains': getattr(self, 'title', None),
            'source__contains': getattr(self, 'source', None),
            'score': getattr(self, 'score', None),
            'areas': getattr(self, 'areas', None),
            'categories': getattr(self, 'categories', None),
            'status': 1,
        }

        args = dict([k, v] for k, v in cond.items() if v)
        queryset = Article.objects.filter(**args)

        c_ids = Category.objects.filter(parent__id=self.category).values_list('id', flat=True)
        if not c_ids:
            queryset = queryset.filter(categories=self.category)
        else:
            queryset = queryset.filter(categories__in=c_ids)

        return queryset.values(*fields).order_by('-pubtime')


class EventsAnalysis(object):

    def __init__(self):
        super(EventsAnalysis, self).__init__()

    # 单一事件
    def getEvent(self, eid):
        fields = ('title', 'socialHarm', 'scope', 'grading', 'pubtime', 'desc')

        queryset = Events.objects.filter(id = eid).values(*fields)

        return queryset

    # 事件信息
    def get_data(self, eid):
        fields = ('title', 'url', 'eventskeyword__name', 'source', 'pubtime', 'sentiment')

        queryset = Article.objects.filter(events__id = eid).values(*fields)

        return queryset

    # 事件内容分析
    def getSentiment(self, eid):
            articles = Article.objects.filter(events__id = eid)

            # 查询情感条数（0 负面， 1 中性， 2 正面）
            sentimentList = []
            s = 3
            for x in range(s):
                sentiment = articles.filter(sentiment = x)

                data = {
                    'sentiment': sentiment.count(),
                }
                sentimentList.append(data)

            return sentimentList

    def getSource(self, eid):
        event = Events.objects.filter(id = eid)
        event = event.annotate(num_source = Count('articles'))
        source = event.values('articles__source', 'num_source').order_by('-num_source')
        
        return source

    def getAreas(self, eid):
        articles = Article.objects.filter(events__id = eid)
        areas = Area.objects.filter(level = 2).values_list('name', flat = True)

        areasList = []
        for a in areas:
            areas_id = Area.objects.get(name = a).id
            areas_sum = articles.filter(Q(areas__id = areas_id) | Q(areas__parent_id = areas_id) | 
                Q(areas__parent__parent_id = areas_id))
            data = {
                'name': a,
                'sum': areas_sum.count(),
            }

            areasList.append(data)

        areasNum = len(areasList)
        for i in range(areasNum):
            if i == range(areasNum):
                continue
            j = i + 1
            while j < areasNum:
                if areasList[i]['sum'] < areasList[j]['sum']:
                    areasListCopy = copy.deepcopy(areasList[j])
                    areasList[j] = copy.deepcopy(areasList[i])
                    areasList[i] = copy.deepcopy(areasListCopy)

                j += 1

        return areasList

    def getKeywords(self, eid):
        event = EventsKeyword.objects.filter(events_id = eid)
        event = event.annotate(num_eventskeyword = Count('articles')).values('name', 'num_eventskeyword')

        return event

    # 事件传播分析
    def getTimeTrend(self, eid):
        event = Events.objects.filter(id = eid)
        time = event.annotate(num_source = Count('articles')).values('articles__pubtime').order_by('articles__pubtime')
        
        return time

    def getTrend(self, eid):
        event = Events.objects.filter(id = eid)
        articles = Article.objects.filter(events__id = eid)

        source = event.annotate(num_source = Count('articles')).values_list('articles__source', flat = True).order_by('-num_source')[:5]
        time = event.annotate(Count('articles')).values_list('articles__pubtime', flat = True).order_by('articles__pubtime')

        data = []
        result = []
        for s in source:
            sourceFilter = articles.filter(source = s)

            timeTrent = []
            for t in time:
                number = sourceFilter.filter(pubtime = t).count()
                timeTrent.append(number)

            data = {
                'name': s,
                'numTrend': timeTrent,
            }

            result.append(data)

        return result 

    # def getWay(self, eid):
    #     event = EventsKeyword.objects.filter(events_id = eid)
    #     eventsmedia = EventsMedia.objects.filter(articles__events__id = eid)
    #     articles = Article.objects.filter(events__id = eid)
        
    #     source = event.annotate(Count('articles')).values_list('articles__source', flat = True)
    #     media = eventsmedia.annotate(Count('eventsmedia')).values('source')
        # print(source)


class RiskData(Abstract):

    def __init__(self, user, params):
        super(RiskData, self).__init__(params)
        self.user = user

    def get_all(self):
        fields = ('id', 'url', 'title', 'source', 'pubtime', 'score', 'status', 'corpus__keyword', 'industry__name',
                 'industry__parent__name', 'harm__id' )

        cond = {
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
            'title__contains': getattr(self, 'title', None),
            'source__contains': getattr(self, 'source', None),
            'score': getattr(self, 'score', None),
            'status': getattr(self, 'status', None),
            'industry_id': getattr(self, 'industry', None),
            'areas': getattr(self, 'areas', None),
            'categories': getattr(self, 'category', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)
        queryset = Article.objects.filter(**args)

        # 判断当前用户是否为武汉深度网科技有限公司成员，然后取出该用户管理的资料
        group_ids = Group.objects.filter(user=self.user).values_list('id', flat=True)
        if 4 in group_ids and 3 in group_ids:
            queryset = queryset.filter(user_id = self.user.id).values(*fields)
        return queryset.values(*fields).order_by('-pubtime')


class RiskDataAdd(Abstract):

    def __init__(self, user, params={}):
        super(RiskDataAdd, self).__init__(params)
        self.user = user

    def add(self):
        url = getattr(self, 'url', '')
        title = getattr(self, 'title', '')
        pubtime = getattr(self, 'pubtime', '')
        score = getattr(self, 'score', 0)
        source = getattr(self, 'source', '')
        areas = getattr(self, 'areas', '')
        categories = getattr(self, 'categories', '')
        industries = getattr(self, 'industries', -1)
        categoriesnum = categories.split(',')

        for x in range(len(categoriesnum)-1):
            print(categories.split(',')[x])
            if  categories.split(',')[x] =='0002' and score == '0':
                score = '1'

        if not pubtime:
            pubtime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if not industries:
            industries = -1

        if not url or not title or not source or not areas or not categories:
            return 400

        if Article.objects.filter(url=url).exists():
            return 202



        # 有多个地域时逗号分隔，并且忽略掉最后一个逗号
        a_ids = areas.split(',')[:-1:]
        area = Area.objects.filter(id__in=a_ids)

        c_ids = categories.split(',')[:-1:]
        category = Category.objects.filter(id__in=c_ids)

        article = Article(
            title=title,
            url=url,
            pubtime=pubtime,
            source=source,
            score=score,
            industry_id=industries,
            user_id=self.user.id,
            status=1,
        )
        article.save()

        article.areas.add(*area)
        article.categories.add(*category)

        article.save()

        return 200


class RiskDataAudit(Abstract):
    def __init__(self, params={}):
        super(RiskDataAudit, self).__init__(params)

    def edit(self, aid):
        audit_ids = aid.split(',')

        Article.objects.filter(id__in=audit_ids).update(status=1)

        return 200


class RiskDataEdit(Abstract):

    def __init__(self, params={}):
        super(RiskDataEdit, self).__init__(params)

    def edit(self, aid):
        edit_id = aid
        title = getattr(self, 'title', '')
        pubtime = getattr(self, 'pubtime', '')
        score = getattr(self, 'score', 0)
        source = getattr(self, 'source', '')
        areas = getattr(self, 'areas', '')
        categories = getattr(self, 'categories', '')

        if not pubtime or not source or not areas or not categories:
            return 400

        article = Article.objects.get(id=edit_id)
        article.title = title
        article.pubtime = pubtime
        article.score = score
        article.source = source
        article.save()

        a_ids = areas.split(',')[:-1:] # 有多个地域时逗号分隔，并且忽略掉最后一个逗号
        c_ids = categories.split(',')[:-1:]

        area = Area.objects.filter(id__in=a_ids)
        category = Category.objects.filter(id__in=c_ids)

        article.areas.clear()
        article.categories.clear()
        article.areas.add(*area)
        article.categories.add(*category)

        article.save()

        return 200


class RiskDataDelete(Abstract):

    def __init__(self, user):
        self.user = user

    def delete(self, aid):
        del_ids = aid.split(',')

        Article.objects.filter(id__in=del_ids).delete()

        return 200


class RiskDataUpload(Abstract):

    def __init__(self, user):
        self.user = user

    def upload(self, filename, file_obj):
        #Model weight
        model = {'标题': 0, 'URL': 0, '发布时间': 0, '来源': 0, '风险程度': 0, '地域': 0, '类别': 0, '行业编号': 0, '监测词': 0}
        #sheet value
        sv = lambda x, y, z : z.cell(row=x, column=y).value
        #date format
        def date_format(df):
            try:
                return openpyxl.utils.datetime.from_excel(df)
            except Exception:
                try:
                    return str_to_date(df)
                except Exception:
                    return df

        try:
            xlsx_book = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
            sheet = xlsx_book.active
            rows = sheet.rows
        except Exception as e:
            return {
                    'status': 0,
                    'message': '操作失败！请检查文件是否有误。详细错误信息：%s！' % e
                }

        total = 0
        dupli = 0

        for i, row in enumerate(rows):
            i += 1
            if i == 1:
                line = [cell.value for cell in row]
                for k in model.keys():
                    model[k] = line.index(k) + 1
            else:
                try:
                    title = str(sv(i, model['标题'], sheet)).strip()
                    if title == 'None':
                        continue

                    url = str(sv(i, model['URL'], sheet)).strip()
                    if url == 'None':
                        continue

                    pubtime = str(date_format(sv(i, model['发布时间'], sheet))).strip()
                    if pubtime == 'None':
                        pubtime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    source = str(sv(i, model['来源'], sheet)).strip()
                    if source == 'None':
                        continue

                    score = str(sv(i, model['风险程度'], sheet)).strip()
                    if score == 'None':
                        continue

                    # 行业编号
                    industry_id = sv(i, model['行业编号'], sheet)
                    if not industry_id or industry_id == '0':
                        industry_id = '-1'

                    # 地域
                    area = str(sv(i, model['地域'], sheet)).strip()
                    if area == 'None':
                        continue
                    else:
                        areas = area.split()
                        a_ids = Area.objects.filter(name__in=areas).values_list('id', flat=True)

                        if len(areas) != len(a_ids):
                            return {
                                'status': 0,
                                'message': '操作失败！请检查第 %s 行 "地域"！' % (i + 1, )
                            }

                        area = Area.objects.filter(id__in=a_ids)

                    # 风险类别
                    category = str(sv(i, model['类别'], sheet)).strip()
                    if category == 'None':
                        continue
                    else:
                        categories = category.split()

                        if categories[0] =='风险快讯' and score =='0':
                            score = '1'

                        c_ids = Category.objects.filter(name__in=categories).values_list('id', flat=True)
                        print(c_ids)
                        if len(categories) != len(c_ids):
                            return {
                                'status': 0,
                                'message': '操作失败！请检查第 %s 行 "类别"！' % (i + 1, )
                            }

                        category = Category.objects.filter(id__in=c_ids)

                    # 监测词
                    monitorWord = str(sv(i, model['监测词'], sheet)).strip()
                    if monitorWord == 'None':
                        monitorWord = ''
                    else:
                        m_id = CorpusCategories.objects.filter(keyword=monitorWord).values_list('id', flat=True)

                        if not m_id.exists():
                            return {
                                'status': 0,
                                'message': '操作失败！请检查第 %s 行 "监测词"！' % (i + 1, )
                            }

                        monitorWord = list(m_id)[0]

                    total += 1

                    # 唯一性
                    old_article = Article.objects.filter(url=url)

                    if old_article.exists():
                        old_article = old_article[0]
                        old_article.title = title
                        old_article.url = url
                        old_article.pubtime = pubtime
                        old_article.source = source
                        old_article.score = score
                        old_article.industry_id = industry_id
                        if not old_article.corpus_id:
                            old_article.corpus_id = monitorWord
                        old_article.save()
                        old_article.areas.clear()
                        old_article.categories.clear()
                        old_article.areas.add(*area)
                        old_article.categories.add(*category)
                        old_article.save()

                        dupli += 1
                        continue

                    article = Article(
                        title=title,
                        url=url,
                        pubtime=pubtime,
                        source=source,
                        score=score,
                        corpus_id=monitorWord,
                        industry_id=industry_id,
                        user_id=self.user.id,
                        status=1,
                        sentiment=-1,
                    )
                    article.save()
                    article.areas.add(*area)
                    article.categories.add(*category)
                    article.save()

                except Exception as e:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
                    }

        return {
                    'status': 1,
                    'message': '操作成功！共处理%s条数据，新增数据%s条，更新数据%s条！' % (total, total - dupli, dupli, )
                }


class RiskDataExport(Abstract):

    def __init__(self, user, params={}):
        self.user = user
        super(RiskDataExport, self).__init__(params)

    def export(self):
        filename = "articles.xlsx"

        # process data
        data = [
            ['标题', 'URL', '发布时间', '来源', '风险程度', '地域', '类别', '行业'],
        ]
        fields = ('id', 'title', 'url', 'pubtime', 'source', 'score', 'areas__name', 'industry__name')
        cond = {
            'areas__id': getattr(self, 'areas', None),
            'status': getattr(self, 'status'),
            'categories__id': getattr(self, 'category', None),
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
        }
        
        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Article.objects.filter(**args).values(*fields)

        # 判断当前用户是否为武汉深度网科技有限公司成员，然后取出该用户管理的资料
        group_ids = Group.objects.filter(user=self.user).values_list('id', flat=True)
        if 4 in group_ids and 3 in group_ids:
            queryset = queryset.filter(user_id = self.user.id).values(*fields)

        for q in queryset:
            industry__name = '无' if q['industry__name'] == 'None' else q['industry__name']
            catogories = categories(q['id'], admin=True, flat=True)
            data.append([q['title'],
                         q['url'],
                         date_format(q['pubtime'], '%Y-%m-%d'),
                         q['source'],
                         q['score'],
                         q['areas__name'],
                         catogories,
                         industry__name,
                        ])

        # write file
        write_by_openpyxl(filename, data, None)

        return open(filename, 'rb')


class RiskDataSuzhou(Abstract):

    def __init__(self, params={}):
        super(RiskDataSuzhou, self).__init__(params)

    def get_risk_data_list(self, search_value):
        fields = ('id', 'title', 'pubtime', 'url', 'source', 'score' )

        args = {}
        ac_id = Category.objects.get(name='风险快讯').id

        if not search_value:
            queryset = Article.objects.filter(categories=ac_id).values(*fields)
        else:
            queryset = Article.objects.filter(Q(title__contains=search_value) | Q(source__contains=search_value)).values(*fields)
            queryset = queryset.filter(categories=ac_id)

        return queryset


class RiskHarmsManageSave(Abstract):

    def __init__(self, params={}):
        super(RiskHarmsManageSave, self).__init__(params)

    def toSave(self):
        hid = getattr(self, 'id')
        cond = {
            'environment': getattr(self, 'modelEn'),
            'activity': getattr(self, 'modelAct'),
            'mind_body': getattr(self, 'modelMab'),
            'behavior': getattr(self, 'modelBeh'),
            'indoor': getattr(self, 'modelTie'),
            'outdoor': getattr(self, 'modelToe'),
            'physics': getattr(self, 'modelPhy'),
            'chemical': getattr(self, 'modelChe'),
            'biology': getattr(self, 'modelBio'),
            'damage_types': getattr(self, 'modelDt'),
            'damage_degree': getattr(self, 'modelTdod'),
            'damage_reason': getattr(self, 'modelDr'),
        }
        age = getattr(self, 'modelAge')
        sex = getattr(self, 'sex')
        # 切片, 切掉第一个和最后一个逗号
        age_ids = age.split(',')[1:-1:]
        sex_ids = sex.split(',')[1:-1:]

        args = dict([k, v] for k, v in cond.items() if v)

        harm = Harm.objects.create(**args)
        harm.save()

        if sex_ids == [] or sex_ids == ['']:
            for age_id in age_ids:
                HarmPeople(
                    age = age_id,
                    sex = 34,
                    harm_id = harm.id,
                ).save()
        else:
            for age_id, sex_id in zip(age_ids, sex_ids):
                HarmPeople(
                    age = age_id,
                    sex = sex_id,
                    harm_id = harm.id,
                ).save()

        article = Article.objects.get(id=hid)
        harm.article = article
        harm.save()

        return 200


class RiskHarmsDetailsData(Abstract):

    def __init__(self, params={}):
        super(RiskHarmsDetailsData, self).__init__(params)

    def get_data(self):
        fields = ('id', 'environment', 'activity', 'mind_body', 'behavior', 'indoor', 'outdoor', 'physics', 'chemical',
                 'biology', 'damage_types', 'damage_degree', 'damage_reason', 'article__industry__name')

        hid = getattr(self, 'id')

        queryset = Harm.objects.filter(article_id=hid).values(*fields)

        return queryset


class RiskHarmsData(Abstract):
    def __init__(self, params={}):
        super(RiskHarmsData, self).__init__(params)

    def get_data(self):
        fields = ('id', 'name', 'desc', 'parent_id')
        cond = {
            'id': getattr(self, 'hid', None),
            'parent_id': getattr(self, 'parent_id', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = HarmIndicator.objects.filter(**args).values(*fields)

        return queryset


class EventsManageData(Abstract):
    def __init__(self, params):
        super(EventsManageData, self).__init__(params)

    def get_data(self):
        # 标题，社会危害程度，影响范围，分级，发布时间，事件总结
        fields = ('id', 'title', 'socialHarm', 'scope', 'grading', 'pubtime', 'desc')
        cond = {
            'title__contains': getattr(self, 'title', None),
            'socialHarm': getattr(self, 'socialHarm', None),
            'scope': getattr(self, 'scope', None),
            'grading': getattr(self, 'grading', None),
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
        }

        args = dict([k, v] for k, v in cond.items() if v)

        queryset = Events.objects.filter(**args).values(*fields)

        return queryset

    def toSave(self):
        title = getattr(self, 'title')
        socialHarm = getattr(self, 'socialHarm')
        scope = getattr(self, 'scope')
        grading = getattr(self, 'grading')
        pubtime = getattr(self, 'pubtime')

        Events(
            title = title,
            socialHarm = socialHarm,
            scope = scope,
            grading = grading,
            pubtime = pubtime,
        ).save()

        return 200

    def toDel(self, eid):
        eids = eid.split(',')

        Events.objects.filter(id__in = eids).delete()

        return 200

    def toUpd(self):
        eid = getattr(self, 'id')
        title = getattr(self, 'title')
        socialHarm = getattr(self, 'socialHarm')
        scope = getattr(self, 'scope')
        grading = getattr(self, 'grading')
        pubtime = getattr(self, 'pubtime')
        desc = getattr(self, 'desc', '')

        event = Events.objects.get(id = eid)
        event.title = title
        event.socialHarm = socialHarm
        event.scope = scope
        event.grading = grading
        event.pubtime = pubtime
        event.desc = desc
        event.save()

        return 200

    # 点击事件名，跳转页面,显示和article关联的数据
    def LinkData(self, eid):
        fields = ('id', 'title', 'url', 'pubtime', 'source', 'eventskeyword__name')
        cond = {
            'title__contains': getattr(self, 'title', None),
            'source__contains': getattr(self, 'source', None),
            'areas__id': getattr(self, 'areas', None),
            'pubtime__gte': getattr(self, 'starttime', None),
            'pubtime__lte': getattr(self, 'endtime', None),
        }
        args = dict([k, v] for k, v in cond.items() if v)

        event_ids = Events.objects.filter(id = eid).values_list('articles__id', flat=True)

        queryset = Article.objects.filter(id__in = event_ids)
        queryset = queryset.filter(**args).values(*fields)

        return queryset


class EventsDataUpload(Abstract):

    def __init__(self, user):
        self.user = user

    def upload(self, filename, file_obj):
        #Model weight
        model = {'标题': 0, 'URL': 0, '发布时间': 0, '来源': 0, '风险程度': 0, '地域': 0, '类别': 0, '行业编号': 0, '关键词': 0, '事件': 0}
        #sheet value
        sv = lambda x, y, z : z.cell(row=x, column=y).value
        #date format
        def date_format(df):
            try:
                return openpyxl.utils.datetime.from_excel(df)
            except Exception:
                try:
                    return str_to_date(df)
                except Exception:
                    return df

        try:
            xlsx_book = openpyxl.load_workbook(BytesIO(file_obj.read()), read_only=True)
            sheet = xlsx_book.active
            rows = sheet.rows
        except Exception as e:
            return {
                    'status': 0,
                    'message': '操作失败！请检查文件是否有误。详细错误信息：%s！' % e
                }

        total = 0
        dupli = 0

        for i, row in enumerate(rows):
            i += 1
            if i == 1:
                line = [cell.value for cell in row]
                for k in model.keys():
                    model[k] = line.index(k) + 1
            else:
                try:
                    title = str(sv(i, model['标题'], sheet)).strip()
                    if title == 'None':
                        continue

                    url = str(sv(i, model['URL'], sheet)).strip()
                    if url == 'None':
                        continue

                    pubtime = str(date_format(sv(i, model['发布时间'], sheet))).strip()
                    if pubtime == 'None':
                        pubtime = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

                    source = str(sv(i, model['来源'], sheet)).strip()
                    if source == 'None':
                        continue

                    score = str(sv(i, model['风险程度'], sheet)).strip()
                    if score == 'None':
                        continue

                    # 行业编号
                    industry_id = sv(i, model['行业编号'], sheet)
                    if not industry_id or industry_id == '0':
                        industry_id = '-1'

                    # 地域
                    area = str(sv(i, model['地域'], sheet)).strip()
                    if area == 'None':
                        continue
                    else:
                        areas = area.split()
                        a_ids = Area.objects.filter(name__in=areas).values_list('id', flat=True)

                        if len(areas) != len(a_ids):
                            return {
                                'status': 0,
                                'message': '操作失败！请检查第 %s 行 "地域"！' % (i + 1, )
                            }

                        area = Area.objects.filter(id__in=a_ids)

                    # 风险类别
                    category = str(sv(i, model['类别'], sheet)).strip()
                    if category == 'None':
                        continue
                    else:
                        categories = category.split()
                        c_ids = Category.objects.filter(name__in=categories).values_list('id', flat=True)

                        if len(categories) != len(c_ids):
                            return {
                                'status': 0,
                                'message': '操作失败！请检查第 %s 行 "类别"！' % (i + 1, )
                            }

                        category = Category.objects.filter(id__in=c_ids)

                    # 关键词
                    keyWord = str(sv(i, model['关键词'], sheet)).strip()
                    if keyWord == 'None':
                        continue
                    else:
                        keyWord_id = EventsKeyword.objects.filter(name = keyWord)
                        if not keyWord_id.exists():
                            EventsKeyword(name = keyWord).save()
                            keyWord_id = EventsKeyword.objects.filter(name = keyWord)

                        keyWord = keyWord_id[0]

                    # 事件
                    event = str(sv(i, model['事件'], sheet)).strip()
                    if event == 'None':
                        continue
                    else:
                        event_id = Events.objects.filter(title = event)

                        if not event_id.exists():
                            return {
                                'status': 0,
                                'message': '操作失败！请检查第 %s 行 "事件"！' % (i + 1, )
                            }

                        event_id = event_id[0]

                    total += 1

                    # 用百度情感分析AI分析标题情感
                    sentiment = BaiDuAI(title)

                    # 唯一性
                    old_article = Article.objects.filter(url=url)

                    if old_article.exists():
                        # 更新新闻
                        old_article = old_article[0]
                        old_article.title = title
                        old_article.url = url
                        old_article.pubtime = pubtime
                        old_article.source = source
                        old_article.score = score
                        old_article.industry_id = industry_id
                        old_article.sentiment = sentiment
                        old_article.save()

                        old_article.areas.clear()
                        old_article.categories.clear()
                        old_article.areas.add(*area)
                        old_article.categories.add(*category)
                        old_article.save()
                        # 关键词和事件和新闻绑定
                        keyWord.events_id = event_id.id
                        keyWord.save()
                        keyWord.articles.add(old_article)
                        keyWord.save()
                        # 事件和新闻绑定
                        event_id.articles.add(old_article)
                        event_id.save()

                        dupli += 1
                        continue

                    article = Article(
                        title=title,
                        url=url,
                        pubtime=pubtime,
                        source=source,
                        score=score,
                        industry_id=industry_id,
                        user_id=self.user.id,
                        status=1,
                        sentiment=sentiment,
                    )
                    article.save()
                    article.areas.add(*area)
                    article.categories.add(*category)
                    article.save()

                    article_id = Article.objects.get(title = article)
                    keyWord.events_id = event_id.id
                    keyWord.save()
                    keyWord.articles.add(article_id)
                    keyWord.save()

                    event_id.articles.add(article_id)
                    event_id.save()

                except Exception as e:
                    return {
                        'status': 0,
                        'message': '操作失败！Excel %s 行存在问题。详细错误信息：%s！' % (i + 1, e)
                    }

        return {
                    'status': 1,
                    'message': '操作成功！共处理%s条数据，新增数据%s条，更新数据%s条！' % (total, total - dupli, dupli, )
                }        


class newsCrawlerData(Abstract):
    def __init__(self, user, params={}):
        super(newsCrawlerData, self).__init__(params)
        self.user = user

    def edit(self):
        word = getattr(self, 'word', '')
        page = getattr(self, 'page', '')
        thread = threading.Thread(target = newsCrawler, args = (word, page))
        thread.start()
        return 200


class StatisticsShow(Abstract):
    def __init__(self, params):
        super(StatisticsShow, self).__init__(params)


    def get_data(self):
        now = datetime.datetime.strptime(datetime.datetime.now().strftime("%Y-%m-%d"), "%Y-%m-%d")
        aWeek = -1-datetime.datetime.now().isoweekday() # 获取从上周六开始计算到今天获得的天数的负数
        today = datetime.datetime.now().date() # 获取今天日期的date类型
        monthInit = datetime.date(today.year, today.month, 1) # 获取本月的月初时间
        aMonth = (monthInit - today).days    #获取今天多少号的负数
        
        time_week = now + datetime.timedelta(days = aWeek)
        time_month = now + datetime.timedelta(days = aMonth)
        time_monthAddweek = now + datetime.timedelta(days = aMonth - 7)

        queryset = Article.objects.filter(pubtime__gte = time_monthAddweek)

        if getattr(self, 'category', None) == '0001' or getattr(self, 'category', None) == '0002':
            category_id = getattr(self, 'category')
            category_ids = Article.objects.filter(categories__id = category_id).values_list('id', flat = True)
            queryset = queryset.filter(id__in = category_ids)
        if getattr(self, 'category', None) == '0003':
            category_business_ids = Article.objects.filter(categories__parent_id = '0003').values_list('id', flat = True)
            queryset = queryset.filter(id__in = category_business_ids)



        user_id = 65
        listdata = []
        while(user_id <= 78):
            if User.objects.filter(id = user_id).exists():
                queryset_short = queryset
                if getattr(self, 'time', None) == '每日':
                    queryset_short = queryset.filter(pubtime__gte = now, user_id = user_id, status=1)
                elif getattr(self, 'time', None) == '每周':
                    queryset_short = queryset.filter(pubtime__gte = time_week, user_id = user_id, status=1)
                elif getattr(self, 'time', None) == '每月':
                    queryset_short = queryset.filter(pubtime__gte = time_month, user_id = user_id, status=1)
                else:
                    queryset_short = queryset.filter(pubtime__gte = time_week, user_id = user_id, status=1)

                data = {
                    'user': user_id,
                    'times': queryset_short.count(),
                }
                listdata.append(data)

            user_id += 1

        return listdata
